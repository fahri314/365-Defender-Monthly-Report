import os
import json
import platform
import requests
import subprocess
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import calendar


class report:
    def __init__(self):
        with open("config.json", "r") as config:
            config = json.loads(config.read())
        # Tenant ID
        self.tenant_ids = {}
        for tenant in config['tenant_ids']:
            self.tenant_ids[tenant['alias']] = tenant['tenant_id']
        # Office Path
        self.excel_path_on_wsl = config['excel_path_on_wsl']
        # TP Settings
        self.exclude_mail_tps = config['exclude_mail_tps']
        self.exclude_benign_positive_alarms = config['exclude_benign_positive_alarms']
        self.exclude_tps_by_keywords = {}
        self.exclude_tps_by_keywords = config['exclude_tps_by_keywords']
        self.tp_incident_links = []
        # Cookie
        self.cookies = {}
        for tenant in config['tenant_ids']:
            self.cookies[tenant['alias']] = tenant['cookie']
        clear_screen()
        self.tenant_id, self.cookie = self.select_tenant()
        clear_screen()
        cookie_keys_to_extract = ['sccauth', 'XSRF-TOKEN', 'ai_session', 's.SessID', 'SSR']
        cookie_values = self.extract_values_from_cookie(self.cookie, cookie_keys_to_extract)
        self.sccauth = cookie_values['sccauth']
        self.xsrf_token = cookie_values['XSRF-TOKEN'].replace('%3A', ":")
        self.ai_session = cookie_values['ai_session']
        self.sess_id = cookie_values['s.SessID']
        self.ssr = cookie_values['SSR']
        # Report date range
        self.from_date, self.to_date = self.get_report_date_range()
        self.edit_list = []
        self.activated = 0
        self.passed = 0
        self.i = 0
        # Get the current directory
        current_dir = os.getcwd()
        # Find all .xlsx files in the current directory
        # Excel file
        self.excel_file = "temp.xlsx"
        # Delete old excel file if exist
        if os.path.exists(self.excel_file):
            os.remove(self.excel_file)
        self.start_time = datetime.now()

    def select_tenant(self):
        # List the available aliases
        print("\nAvailable Tenants:\n")
        for i, alias in enumerate(self.tenant_ids.keys(), start=1):
            print(f"{i}. {alias}")

        # Select a tenant ID by alias
        while True:
            alias_input = input("\nEnter the number of the desired alias: ")
            try:
                alias_num = int(alias_input)
                if 1 <= alias_num <= len(self.tenant_ids):
                    selected_alias = list(self.tenant_ids.keys())[alias_num - 1]
                    selected_cookie = list(self.cookies.keys())[alias_num - 1]
                    return self.tenant_ids[selected_alias], self.cookies[selected_cookie]
                else:
                    print("Invalid input. Please try again.")
                    exit()
            except ValueError:
                print("Invalid input. Please try again.")
                exit()

    def get_report_start_time(self, relative_time):
        first_day_of_month = datetime(relative_time.year, relative_time.month, 1)
        first_day_start = datetime.combine(first_day_of_month, datetime.min.time())
        first_day_start += timedelta(minutes=1)  # İlk dakikayı ekleyerek başlangıç zamanı ayarla
        
        return first_day_start.strftime("%Y-%m-%dT%H:%M:%S.000Z")
    
    def get_report_end_time(self, relative_time):
        last_day_of_month = calendar.monthrange(relative_time.year, relative_time.month)[1]
        last_day = datetime(relative_time.year, relative_time.month, last_day_of_month, 23, 59, 59)
        
        return last_day.strftime("%Y-%m-%dT%H:%M:%S.000Z")

    def get_report_date_range(self):
        today = datetime.today()
        response = input("Are you within the report date range? [y/N] ").strip() or 'n'

        if response.lower() == 'y':
            from_date = self.get_report_start_time(today)
            to_date = self.get_report_end_time(today)
        else:
            last_month = today.month - 1 if today.month > 1 else 12
            last_year = today.year if today.month > 1 else today.year - 1
            last_day_of_last_month = calendar.monthrange(last_year, last_month)[1]
            last_month_end = datetime(last_year, last_month, last_day_of_last_month)
            
            from_date = self.get_report_start_time(last_month_end)
            to_date = self.get_report_end_time(last_month_end)

        return from_date, to_date

    def get_incidents(self):
        incidents = []
        page_index = 0
        uri = "https://security.microsoft.com/apiproxy/mtp/incidentQueue/incidents/alerts"
        headers, cookies = self.generate_header_data()
        print("[+] Incidents Downloading...")
        # This method is working like 'Export' download button. Max page size is 50
        while True:
            page_index += 1
            print("page_index:", page_index)
            post_data = self.generate_post_data(page_index)
            response = requests.post(uri, json = post_data, headers = headers, cookies = cookies)
            if response.text == '[]':
                break
            if response.status_code != 200:
                print("Response: ", response.text)
                raise Exception("Unable to get incidents from tenant, did the session time out?")
            incidents.extend(json.loads(response.text))

        return incidents

    def get_devices(self):
        devices = []
        page_index = 0
        headers, cookies = self.generate_header_data()
        print("[+] Devices Downloading...")
        # This method is working like 'Export' download button. Max tested page size is 200
        while True:
            page_index += 1
            uri = f"https://security.microsoft.com/apiproxy/mtp/k8s/machines?tid={self.tenant_id}&deviceCategories=Endpoint&onBoardingStatuses=Onboarded&lookingBackIndays=30&pageIndex={page_index}&pageSize=200"
            print("page_index:", page_index)
            response = requests.get(uri, headers = headers, cookies = cookies)
            if response.text == '[]':
                break
            devices.extend(json.loads(response.text))
            if response.status_code != 200:
                print("Response: ", response.text)
                raise Exception("Unable to get devices from tenant, did the session time out?")

        return devices

    def get_iocs(self, ioc_type):
        iocs = []
        uri = f"https://security.microsoft.com/apiproxy/mtp/papin/api/cloud/public/internal/indicators/getQuery?type={ioc_type}&PageIndex=0&PageSize=99999&tid={self.tenant_id}"
        headers, cookies = self.generate_header_data()
        response = requests.get(uri, headers = headers, cookies = cookies)
        iocs.extend(json.loads(response.text))
        if response.status_code != 200:
            print("Response: ", response.text)
            raise Exception(f"Unable to get {ioc_type} from tenant, did the session time out?")

        return iocs

    def group_severities(self, incidents):
        # Create a default dictionary to hold groups
        severity_groups = defaultdict(list)

        # Group incidents by severity for the target classification
        for incident in incidents:
            severity = incident.get("Severity", "Unspecified")
            severity_groups[severity].append(incident)

        return severity_groups

    def group_severities_by_classification(self, incidents, target_classification):
        # Create a default dictionary to hold groups
        severity_groups = defaultdict(list)

        # Group incidents by severity for the target classification
        for incident in incidents:
            if incident.get("Classification") == target_classification:
                severity = incident.get("Severity", "Unspecified")
                severity_groups[severity].append(incident)

        return severity_groups
    
    def group_resolves(self, incidents):
        # Create a default dictionary to hold groups
        resolve_groups = defaultdict(list)

        # Group incidents by severity for the target classification
        for incident in incidents:
            classification = incident.get("Classification", "Unspecified")
            resolve_groups[classification].append(incident)

        return resolve_groups
    
    def group_resolves_by_severity(self, incidents, target_severity):
        # Create a default dictionary to hold groups
        resolve_groups = defaultdict(list)

        # Group incidents by severity for the target severity
        for incident in incidents:
            if incident.get("Severity") == target_severity:
                classification = incident.get("Classification", "Unspecified")
                resolve_groups[classification].append(incident)

        return resolve_groups

    def group_incident_categories(self, incidents): 
        # Create a default dictionary to hold groups
        category_groups = defaultdict(list)

        # Group category
        for incident in incidents:
            if self.exclude_benign_positive_alarms:
                if incident.get("Classification") == classifications["Benign Positive"]:
                    continue
            category = incident.get("Categories", "Unspecified")[0]
            category_groups[category].append(incident)

        return category_groups

    def group_device_oses(self, devices): 
        # Create a default dictionary to hold groups
        os_groups = defaultdict(list)

        # Group devices os
        for device in devices:
            os = device.get("OsPlatform", "Unspecified")
            if os == None:
                continue
            os_groups[os].append(device)

        return os_groups

    def group_device_os_types(self, devices):
        os_groups = {
            "Windows Client": [],
            "Windows Server": [],
            "Linux": [],
            "Android": [],
            "macOS": [],
            "iOS": [],
            "Other": []
        }

        for device in devices:
            os = device.get("OsPlatform", "Unspecified")
            if os == None:
                continue
            if "Server" in os:
                os_groups["Windows Server"].append(device)
            elif "WindowsXP" in os:
                os_groups["Windows Client"].append(device)
            elif "Windows7" in os:
                os_groups["Windows Client"].append(device)
            elif "Windows8" in os:
                os_groups["Windows Client"].append(device)
            elif "Windows10" in os:
                os_groups["Windows Client"].append(device)
            elif "Windows11" in os:
                os_groups["Windows Client"].append(device)
            elif "Windows12" in os:
                os_groups["Windows Client"].append(device)
            elif "Linux" in os:
                os_groups["Linux"].append(device)
            elif "Android" in os:
                os_groups["Android"].append(device)
            elif "macOS" in os:
                os_groups["macOS"].append(device)
            elif "iOS" in os:
                os_groups["iOS"].append(device)
            else:
                os_groups["Other"].append(device)
        return os_groups

    def group_incident_sources(self, incidents): 
        # Create a default dictionary to hold groups
        incident_source_groups = defaultdict(list)
        incident_source_short_names = {
            "Microsoft Defender for Office 365": "Office 365",
            "Microsoft Defender for Endpoint": "Endpoint",
            "Microsoft Defender for Cloud Apps": "Cloud Apps",
            "Microsoft Defender XDR": "Defender XDR"
        }

        # Group incident sources
        for incident in incidents:
            if self.exclude_benign_positive_alarms:
                if incident.get("Classification") == classifications["Benign Positive"]:
                    continue
            source = incident.get("ProductNames", "Unspecified")[0]
            if source in incident_source_short_names:
                source = incident_source_short_names[source]
            incident_source_groups[source].append(incident)

        return incident_source_groups

    def detect_impacted_entities(self, data):
        # Get the impacted entities dictionary
        impacted_entities = data.get("ImpactedEntities", {})

        # Get ComputerDnsName values from the Machines list
        machines = impacted_entities.get("Machines", [])
        if machines:
            computer_dns_names = [machine.get("ComputerDnsName") for machine in machines if machine.get("ComputerDnsName")]
            if computer_dns_names:
                return ", ".join(computer_dns_names)

        # If Machines list is empty, get UserName values from the Users list
        users = impacted_entities.get("Users", [])
        if users:
            user_names = [user.get("UserName") for user in users if user.get("UserName")]
            if user_names:
                return ", ".join(user_names)

        # If Users list is also empty, get DisplayName values from the Mailboxes list
        mailboxes = impacted_entities.get("Mailboxes", [])
        if mailboxes:
            display_names = [mailbox.get("DisplayName") for mailbox in mailboxes if mailbox.get("DisplayName")]
            if display_names:
                return ", ".join(display_names)

        # If all lists are empty, return an empty string
        return ""
    
    def get_analyst_comment(self, incident_id):
        headers, cookies = self.generate_header_data()
        uri = f"https://security.microsoft.com/apiproxy/mtp/auditHistory/AuditHistory?&entityType=IncidentEntity&id={incident_id}&auditType=0&pageIndex=1&pageSize=100"
        response = requests.get(uri, headers = headers, cookies = cookies)
        if response.status_code != 200:
            print("Response: ", response.text)
            raise Exception("Unable to get the audit history from tenant, did the session time out?")

        data = response.json()
        for item in data:
            if item.get("type") == "Feedback":
                return item.get("newValue")
        return None

    def get_total_mail_incident_count(self, incidents):
        count = 0
        for incident in incidents:
            incident_name = incident.get("Title")
            detection_source = incident.get("DetectionSources")
            if detection_source[0] == detection_sources["Office 365"] or "Maalware Baazar" in incident_name:
                count += 1
        return count

    def get_tp_incident_count(self, incidents):
        count = 0
        for incident in incidents:
            if incident.get("Classification") != classifications["True Positive"]:
                continue  # Skip incidents that are not TP IDs
            count += 1
        return count

    def get_tp_mail_incident_count(self, incidents):
        count = 0
        for incident in incidents:
            if incident.get("Classification") == classifications["True Positive"]:
                incident_name = incident.get("Title")
                detection_source = incident.get("DetectionSources")
                if detection_source[0] == detection_sources["Office 365"] or "Maalware Baazar" in incident_name:
                    count += 1
        return count

    def print_resolve_dist(self, incidents):
        # Group incidents by resolution classification
        grouped_resolves = self.group_resolves(incidents)
        # Sort classifications based on the number of incidents in descending order
        sorted_resolves = sorted(
            grouped_resolves.items(), key=lambda item: len(item[1]), reverse=True
        )
        resolve_dist_table_data = ""
        print(f"\n\x1b[1;31;43m[+] Resolve Distribution\x1b[0;0m\n")
        # Iterate through sorted classifications and print the results
        for classification, incidents in sorted_resolves:
            if classification != 0:
                resolve_dist_table_data += f"{classifications_reverse[classification]}\t{len(incidents)}\n"
                print(f"{classifications_reverse[classification]}\t{len(incidents)}")
        self.copy_to_excel(resolve_dist_table_data, "Resolve Dist")

    def print_severity_dist(self, incidents):
        grouped_severities = self.group_severities(incidents)
        # Sort severities based on the number of incidents in descending order
        sorted_severities = sorted(
            grouped_severities.items(), key=lambda item: len(item[1]), reverse=True
        )
        severity_dist_table_data = ""
        print(f"\n\x1b[1;31;43m[+] Severity Distribution\x1b[0;0m\n")
        for severity, incidents in sorted_severities:
            severity_dist_table_data += f"{severities[severity]}\t{len(incidents)}\n"
            print(f"{severities[severity]}\t{len(incidents)}")
        self.copy_to_excel(severity_dist_table_data, "Severity Dist")

    def print_tp_severity_dist(self, incidents):
        # Group incidents by severity for the "True Positive" classification
        grouped_severities = self.group_severities_by_classification(incidents, classifications["True Positive"])
        # Sort severities based on the number of incidents in descending order
        sorted_severities = sorted(
            grouped_severities.items(), key=lambda item: len(item[1]), reverse=True
        )
        print(f"\n\x1b[1;31;43m[+] TP Severity Distribution\x1b[0;0m\n")
        # Iterate through sorted severities and print the results
        for severity, incidents in sorted_severities:
            print(f"{severities[severity]}\t{len(incidents)}")
    
    def print_high_severity_resolve_dist(self, incidents):
        # Group incidents by resolution classification for "High" severity
        grouped_resolves = self.group_resolves_by_severity(incidents, severities_reverse["High"])
        # Sort classifications based on the number of incidents in descending order
        sorted_resolves = sorted(
            grouped_resolves.items(), key=lambda item: len(item[1]), reverse=True
        )
        print(f"\n\x1b[1;31;43m[+] High Severity Resolve Distribution\x1b[0;0m\n")
        resolve_dist_table_data = ""
        # Iterate through sorted classifications and print the results
        for classification, incidents in sorted_resolves:
            if classification != 0:
                resolve_dist_table_data += f"{classifications_reverse[classification]}\t{len(incidents)}\n"
                print(f"{classifications_reverse[classification]}\t{len(incidents)}")
        # Save sorted resolve distribution data to Excel
        self.copy_to_excel(resolve_dist_table_data, "High Severity Resolve Dist")

    def print_incidents_category_dist(self, incidents):
        grouped_incident_categories = self.group_incident_categories(incidents)
        print(f"\n\x1b[1;31;43m[+] Incidents Category Distribution\x1b[0;0m\n")
        for category, incidents in grouped_incident_categories.items():
            print(f"{category}\t{len(incidents)}")

    def print_device_os_dist(self, devices):
        # Group devices by OS
        grouped_device_oses = self.group_device_oses(devices)
        # Sort OS groups based on the number of devices in descending order
        sorted_device_oses = sorted(
            grouped_device_oses.items(), key=lambda item: len(item[1]), reverse=True
        )
        print(f"\n\x1b[1;31;43m[+] Device OS Distribution\x1b[0;0m\n")
        device_os_dist_table_data = ""
        # Iterate through sorted OS groups and print the results
        for os_name, devices in sorted_device_oses:
            device_os_dist_table_data += f"{os_name}\t{len(devices)}\n"
            print(f"{os_name}\t{len(devices)}")
        # Save sorted device OS distribution data to Excel
        self.copy_to_excel(device_os_dist_table_data, "OS Dist")
    
    def print_device_os_type_dist(self, devices):
        # Group devices by OS type
        grouped_device_os_types = self.group_device_os_types(devices)
        # Sort OS types based on the number of devices in descending order
        sorted_device_os_types = sorted(
            grouped_device_os_types.items(), key=lambda item: len(item[1]), reverse=True
        )
        print(f"\n\x1b[1;31;43m[+] Device OS Type Distribution\x1b[0;0m\n")
        device_os_type_dist_table_data = ""
        # Iterate through sorted OS types and print the results
        for os_type, os in sorted_device_os_types:
            if len(os) != 0:
                device_os_type_dist_table_data += f"{os_type}\t{len(os)}\n"
                print(f"{os_type}\t{len(os)}")
        # Save sorted device OS type distribution data to Excel
        self.copy_to_excel(device_os_type_dist_table_data, "OS Type Dist")

    def print_tp_incidents_and_comments(self, incidents):
        print(f"\n\x1b[1;31;43m[+] TP Incidents Table\x1b[0;0m\n")
        tp_incident_id_list = []
        tp_incident_table_data = ""

        for incident in incidents:
            classification = incident.get("Classification")
            if classification != classifications["True Positive"]:
                continue  # Skip incidents that are not TP IDs

            incident_source = incident.get("ProductNames")[0]
            incident_name = incident.get("Title")

            # Exclude conditions
            if self.exclude_mail_tps:
                if incident_source in ["Microsoft Defender XDR", "Microsoft Defender for Office 365"]:
                    continue  # Skip if incident source is excluded
                if self.contains_any(incident_name, *self.exclude_tps_by_keywords):
                    continue  # Skip if incident name contains excluded keywords
            else:
                if self.contains_any(incident_name, *self.exclude_tps_by_keywords):
                    continue  # Skip if incident name contains excluded keywords

            last_activity = incident.get("LastUpdateTime")
            severity = severities.get(incident.get("Severity"), "Unknown")
            impacted_entities = self.detect_impacted_entities(incident)
            incident_id = incident.get("IncidentId")

            # Format last activity time
            last_activity = last_activity.rstrip('Z')  # Remove the 'Z' character
            last_activity = last_activity.split('.')[0]  # Remove the microsecond part
            last_activity = datetime.strptime(last_activity, "%Y-%m-%dT%H:%M:%S").strftime("%d-%m-%Y - %H:%M")

            incident_url = f"https://security.microsoft.com/incident2/{incident_id}/overview?tid={self.tenant_id}"
            self.tp_incident_links.append(incident_url)
            tp_incident_table_data += f"{incident_id}\t{last_activity}\t{severity}\t{incident_name}\tTrue Positive\tFile Block​\t{impacted_entities}\n"
            # Print TP incident details
            incident_id_link = self.clickable_link(incident_id, incident_url)
            print(f"{incident_id_link}\t{last_activity}\t{severity}\t{incident_name}\tTrue Positive\tFile Block​\t{impacted_entities}")

            tp_incident_id_list.append(incident_id)

        # Save TP incident table to excel
        self.copy_to_excel(tp_incident_table_data, "TP Incidents")

        print(f"\n\x1b[1;31;43m[+] TP Incidents Analyst Feedback Comments\x1b[0;0m\n")
        # Print Analyst Feedback Comments
        for incident_id in tp_incident_id_list:
            analyst_comment = self.get_analyst_comment(incident_id)
            print(f"• ({incident_id}) {analyst_comment}")

    def open_tp_incidents(self):
        response = input("Do you want to open the TP incidents on the browser? [Y/n] ").strip() or 'y'

        if response.lower() == 'y':
            for url in report.tp_incident_links:
                self.open_url(url)

    def print_incident_source_dist(self, incidents):
        grouped_incident_sources = self.group_incident_sources(incidents)
        print(f"\n\x1b[1;31;43m[+] Incident Sources Distribution\x1b[0;0m\n")
        for source, incidents in grouped_incident_sources.items():
            print(f"{source}\t{len(incidents)}")

    def generate_header_data(self):
        headers = {
            "authority": "security.microsoft.com",
            "method": "POST",
            "path": f"/apiproxy/mtp/huntingService/rules?tenantIds[]={self.tenant_id}",
            "scheme": "https",
            "accept": "application/json, text/plain, */*",
            "accept-encoding": "gzip, deflate, br, zstd",
            "accept-language": "tr-tr",
            "m-connection": "4g",
            "m-viewid": "",
            "origin": "https://security.microsoft.com",
            "priority": "u=1, i",
            "referer": "https://security.microsoft.com/v2/advanced-hunting?tid={self.tenant_id}",
            "sec-ch-ua": '"Not)A;Brand";v="99", "Google Chrome";v="127", "Chromium";v="127"',
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": '"Windows"',
            "sec-fetch-dest": "empty",
            "sec-fetch-mode": "cors",
            "sec-fetch-site": "same-origin",
            "tenant-id": self.tenant_id,
            "x-accepted-statuscode": "3..|4..|50.",
            "x-clientpage": "hunting-2@wicd-hunting",
            "x-tabvisible": "visible",
            "x-tid": self.tenant_id,
            "x-xsrf-token": self.xsrf_token
        }

        cookies = {
            "SSR": self.ssr,
            "at_check": "true",
            "BCP": "AD=1&AL=1&SM=1",
            "SRCHHPGUSR": "SRCHLANG=tr&DM=1&PV=15.0.0&CIBV=1.1418.9-suno",
            "i18next": "tr-TR",
            "s.SessID": self.sess_id,
            "s.Flight": "",
            "sccauth": self.sccauth,
            "X-PortalEndpoint-RouteKey": "neuprod_northeurope",
            "XSRF-TOKEN": self.xsrf_token,
            "ai_session": self.ai_session
        }

        return headers, cookies

    def generate_post_data(self, page_index):
        return {
            "isDexLicense": False,
            "isStatusFilterEnable": False,
            "isUSXIncidentAssignmentEnabled": True,
            "pageSize": 50,
            "isMultipleIncidents": True,
            "serviceSources": {
                "1": [
                    "AutomatedInvestigation",
                    "CustomDetection",
                    "MTP",
                    "CustomerTI",
                    "Bitdefender,Ziften,SentinelOne,Lookout",
                    "WindowsDefenderSmartScreen",
                    "WindowsDefenderAv",
                    "WindowsDefenderAtp"
                ],
                "2": [
                    "8192"
                ],
                "4": [
                    "16384"
                ],
                "8": [
                    "OfficeATP"
                ],
                "16": [
                    "CustomDetection",
                    "MTP",
                    "Manual"
                ],
                "32": [
                    "AAD"
                ],
                "64": [
                    "AppGPolicy",
                    "AppGDetection"
                ]
            },
            "fromDate": self.from_date,
            "toDate": self.to_date,
            "pageIndex": page_index,
            "sortOrder": "Descending",
            "sortByField": "LastUpdateTime"
        }

    def extract_values_from_cookie(self, cookie, keys):
        # Split the cookie string into individual key-value pairs
        cookie_pairs = cookie.split('; ')
        # Convert to dictionary for easy access
        cookie_dict = {pair.split('=')[0]: pair.split('=')[1] for pair in cookie_pairs}
        # Extract the desired values
        extracted_values = {key: cookie_dict.get(key) for key in keys}
        return extracted_values

    def clickable_link(self, text, url):
        # ANSI escape code for clickable link
        clickable_link = f"\033]8;;{url}\033\\{text}\033]8;;\033\\"
        return clickable_link

    def open_url(self, url):
        print(url)
        if os.name == 'posix':
            if 'microsoft' in platform.uname().release.lower():
                os.system(f'wslview {url}')
            else:
                os.system(f'xdg-open {url}')
        else:
            os.system(f'start {url}')

    def copy_to_excel(self, data, sheet_name):
        # Check if the file exists, open it; otherwise, create a new one
        if os.path.exists(self.excel_file):
            wb = load_workbook(self.excel_file)
        else:
            wb = Workbook()
            # Remove the default sheet created by Workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)
    
        # If the sheet already exists, remove it first to avoid duplication
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
    
        # Create a new sheet
        ws = wb.create_sheet(title=sheet_name)
    
        # Split the data into rows and append to the worksheet
        rows = data.split('\n')
        for row in rows:
            values = row.split('\t')
            formatted_values = []
            for value in values:
                # Try to convert to a number if possible
                try:
                    formatted_value = float(value) if '.' in value else int(value)
                except ValueError:
                    formatted_value = value
                formatted_values.append(formatted_value)
            ws.append(formatted_values)
    
        # Save and close the workbook
        wb.save(self.excel_file)
        wb.close()

    def open_excel(self):
        excel_process = None  # Store process reference
        if 'microsoft' in platform.uname().release.lower():  # WSL
            try:
                temp_file_win_path = subprocess.check_output(['wslpath', '-w', self.excel_file]).decode().strip()
                excel_process = subprocess.Popen([self.excel_path_on_wsl, temp_file_win_path])
            except subprocess.CalledProcessError:
                print("Excel not found on Windows. Please ensure Excel is installed.")
        elif platform.system() == 'Windows':
            try:
                # Open Excel and store the process reference
                excel_process = subprocess.Popen(['start', 'excel', os.path.abspath(self.excel_file)], shell=True)
            except Exception as e:
                print(f"Failed to open Excel: {e}")
        else:
            print("This operation is only supported in WSL or Windows environments.")
            return

    def contains_any(self, incident_name, *texts):
        return any(text in incident_name for text in texts)

severities = {
    32: 'Informational',
    64: 'Low',
    128: 'Medium',
    256: 'High',
}

severities_reverse = {
    'Informational': 32,
    'Low': 64,
    'Medium': 128,
    'High': 256,
}

classifications = {
    'False Positive': 10,
    'True Positive': 20,
    'Benign Positive': 30 
}

classifications_reverse = {
    10 : 'False Positive',
    20 : 'True Positive',
    30 : 'Benign Positive'
}

detection_sources = {
    'Office 365': 512,      # Microsoft Defender for Office 365
    'Defender XDR': 4096    # Microsoft Defender XDR
}

def clear_screen():
    os.system('cls' if os.name == 'nt' else 'clear')

if __name__ == '__main__':
    report = report()
    clear_screen()
    incidents = report.get_incidents()
    devices = report.get_devices()
    domains_and_urls = report.get_iocs("url")
    file_hashes = report.get_iocs("files")
    ips = report.get_iocs("ip")

    clear_screen()

    # Report Date Range
    print("\n\x1b[1;31;43m[+] Report date range\x1b[0;0m\n")
    from_date = datetime.strptime(report.from_date, "%Y-%m-%dT%H:%M:%S.%fZ").strftime("%d.%m.%y")
    to_date = datetime.strptime(report.to_date, "%Y-%m-%dT%H:%M:%S.%fZ").strftime("%d.%m.%y")
    print(from_date, "–", to_date)

    # Total number of devices
    total_device_count = len(devices)
    print(f"\n\x1b[1;31;43m[+] Total Device\x1b[0;0m\n")
    print(total_device_count)

    # Total Incident
    total_incident = len(incidents)
    total_mail_incident_count = report.get_total_mail_incident_count(incidents)
    tp_incident_count = report.get_tp_incident_count(incidents)
    tp_mail_incident_count = report.get_tp_mail_incident_count(incidents)
    print(f"\n\x1b[1;31;43m[+] Total Incident\x1b[0;0m\n")
    print(f"TP: {tp_incident_count} - TP Mail: {tp_mail_incident_count}")
    print(f"Total: {total_incident} - Total Mail: {total_mail_incident_count}")

    # OS Type Distribution
    report.print_device_os_type_dist(devices)

    # OS Distribution (Onboarding Endpoint OS Distribution)
    report.print_device_os_dist(devices)

    # Resolve Distribution
    report.print_resolve_dist(incidents)

    # Severity Distribution
    report.print_severity_dist(incidents)

    # High Severity Resolve Distribution
    report.print_high_severity_resolve_dist(incidents)
    
    # # TP Severity Distribution
    # report.print_tp_severity_dist(incidents)

    # # Incidents Category Distribution
    # report.print_incidents_category_dist(incidents)

    # TP Incidents and Comments
    report.print_tp_incidents_and_comments(incidents)

    # # Incident Source Distribution
    # report.print_incident_source_dist(incidents)

    # IOC Counts
    print(f"\n\x1b[1;31;43m[+] Total IOCs\x1b[0;0m\n")
    print("Domain/URL Count:\t", len(domains_and_urls))
    print("File Hash Count:\t", len(file_hashes))
    print("IP Count:\t\t", len(ips))

    end_time = datetime.now()
    print("\n\n\x1b[1;31;43m[!] Elapsed time: ", end_time - report.start_time, "\x1b[0;0m\n")
    print("==================================================")

    # Open excel file
    report.open_excel()
    # Open TP Links on Browser
    report.open_tp_incidents()
